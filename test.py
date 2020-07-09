# coding=UTF-8
import openpyxl

#wb = openpyxl.Workbook()    #创建工作簿
#ws = wb.active  #创建工作表
#ws.title = "newsheet"
#ws1 = wb.create_sheet('sheet2')

wb = openpyxl.load_workbook('d:/test/123456.xlsx')   #加载存在的excel文件
ws = wb.get_sheet_by_name('newsheet')
for x in ws.rows:
    for cell in x:

        print(cell.value)
#ws.cell(row=1,column=1,value='1:1')  向指定行列覆盖字符

#print(ws.cell(row=1,column=1).value)
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

#wb.save('D:/test/123