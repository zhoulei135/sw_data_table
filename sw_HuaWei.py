# -- coding: UTF-8 --
from openpyxl import load_workbook, Workbook
import re

path = 'D:/HuaweiDriver/PyProject/SW/'
path_log = path + 'sw.log'
path_out = path + 'out.xlsx'

fo = open(path_log, 'r', encoding="UTF-8")
wb = Workbook()
ws = wb.active

svi = {}
intf = {}  # [带宽，描述，接口类型，vlan，聚合，ip，shudown]
int_status = {}
int_flag = 0
interface = ''
for line in fo:
    if re.search(r'^interface Vlanif', line):
        # print(line)
        int_flag = 1
        vlan = line.replace('interface Vlanif', '').replace('\n', '').replace(' ', '')
        svi[vlan] = ''
    elif re.search(r'^interface (X)*GigabitEthernet|^interface HundredGig', line):
        # print(line)
        int_flag = 2
        if 'HundredGigE' in line:
            # print(line)
            interface = line.replace('interface HundredGigE', 'HGE')
            interface = interface.replace('\n', '')
            intf[interface] = [100, '', '', '', '', '', '']
        if 'XGigabitEthernet' in line:
            # print(line)
            interface = line.replace('interface XGigabitEthernet', 'XGE')
            interface = interface.replace('\n', '')
            interface = interface.replace(' ', '')
            intf[interface] = [10, '', '', '', '', '', '']
        if ' GigabitEthernet' in line:
            # print(line)
            interface = line.replace('interface GigabitEthernet', 'GE')
            interface = interface.replace('\n', '')
            interface = interface.replace(' ', '')
            intf[interface] = [1, '', '', '', '', '', '']
    elif 'InUti/OutUti' in line:
        int_flag = 3
    elif 'ip address' in line and int_flag == 1:
        ip = line.split()[2]
        svi[vlan] = ip
    elif ' description' in line and int_flag == 2:
        intf[interface][1] = line.replace(' description ', '').replace('\n', '')
    elif 'port link-type trunk' in line and int_flag == 2:
        intf[interface][2] = 'trunk'
    elif re.search(r'^ port trunk allow-pass vlan', line) and int_flag == 2:
        # print(line)
        intf[interface][3] = line.replace(' port trunk allow-pass vlan ', '').replace('\n', '')
        # print(intf[interface])
    elif 'port default vlan' in line and int_flag == 2:
        intf[interface][2] = 'access'
        vid = line.replace(' port default vlan ', '').replace('\n', '').replace(' ', '')
        intf[interface][3] = int(vid)
        intf[interface][5] = svi[vid]

    elif 'eth-trunk' in line and int_flag == 2:
        intf[interface][4] = line.replace(' eth-trunk ', 'eth-trunk').replace('\n', '')
    elif 'shutdown' in line and int_flag == 2:
        intf[interface][6] = 'ADM'
    elif re.search(r'GigabitEthernet\d', line) and int_flag == 3:
        str = line.replace('GigabitEthernet', 'GE')
        str = str.split()
        int_status[str[0]] = ['', '', '空闲']
        if str[1] == '*down':
            # print(str)
            int_status[str[0]][0] = 'ADM'
        if str[1] != '*down':
            #print(str)
            if str[1] == 'up':
                int_status[str[0]][0] = 'up'
            else:
                int_status[str[0]][0] = 'undo shutdown'
        if str[1] == 'up':
            int_status[str[0]][2] = '占用'
    elif '#' in line and int_flag == 1:
        int_flag = 0

i = 1
for k, v in int_status.items():
    ws.cell(row=i, column=4).value = k
    ws.cell(row=i, column=5).value = intf[k][4]
    ws.cell(row=i, column=6).value = intf[k][2]
    ws.cell(row=i, column=7).value = intf[k][3]
    ws.cell(row=i, column=8).value = intf[k][0]
    ws.cell(row=i, column=9).value = v[2]
    ws.cell(row=i, column=10).value = intf[k][5]
    ws.cell(row=i, column=12).value = intf[k][1]
    ws.cell(row=i, column=19).value = v[0]
    # ws.cell(row=i, column=19).value = v[1]
    i += 1

print(svi)
print(intf)
print(int_status)

fo.close()
wb.save(path_out)
wb.close()
