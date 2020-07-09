# -- coding: UTF-8 --
#by liaohongcheng
from openpyxl import load_workbook, Workbook
import re
import os

LogPath='D:/test/'
path_out = LogPath + 'out.xlsx'

def getFullLogPath():
    FullLogPath=[]
    LogName=os.listdir(LogPath)
    for x in LogName:
        if os.path.splitext(x)[1] == '.log':
            FullLogPath.append(LogPath+x)
    return FullLogPath


def outPutToexcle(i,int_status,intf,ws,HostName):
    for k, v in int_status.items():
        ws.cell(row=i, column=1).value = HostName
        ws.cell(row=i, column=4).value = k
        ws.cell(row=i, column=5).value = intf[k][4]
        ws.cell(row=i, column=6).value = intf[k][2]
        ws.cell(row=i, column=7).value = intf[k][3]
        ws.cell(row=i, column=8).value = intf[k][0]
        ws.cell(row=i, column=9).value = v[2]
        ws.cell(row=i, column=10).value = intf[k][5]
        ws.cell(row=i, column=12).value = intf[k][1]
        ws.cell(row=i, column=19).value = v[0]
        # ws.cell(row=i, column=19).value = v[1]
        i += 1
    return i



def DataProcess(Fo):
        intf = {}  # [带宽，描述，接口类型，vlan，聚合，ip，shudown]
        int_status = {}
        svi = {}
        int_flag = 0
        interface = ''
        for Line in Fo:
            if re.search(r'^interface Vlanif', Line):
                int_flag = 1
                vlan = Line.replace('interface Vlanif', '').replace('\n', '').replace(' ', '')
                svi[vlan] = ''
            elif re.search(r'^interface (X)*GigabitEthernet|^interface HundredGig', Line):
                int_flag = 2
                if 'HundredGigE' in Line:
                    interface = Line.replace('interface HundredGigE', 'HGE')
                    interface = interface.replace('\n', '')
                    intf[interface] = [100, '', '', '', '', '', '']
                if 'XGigabitEthernet' in Line:
                    interface = Line.replace('interface XGigabitEthernet', 'XGE')
                    interface = interface.replace('\n', '')
                    interface = interface.replace(' ', '')
                    intf[interface] = [10, '', '', '', '', '', '']
                if ' GigabitEthernet' in Line:
                    interface = Line.replace('interface GigabitEthernet', 'GE')
                    interface = interface.replace('\n', '')
                    interface = interface.replace(' ', '')
                    intf[interface] = [1, '', '', '', '', '', '']
            elif 'InUti/OutUti' in Line:
                int_flag = 3
            elif 'ip address' in Line and int_flag == 1:
                ip = Line.split()[2]
                svi[vlan] = ip
            elif ' description' in Line and int_flag == 2:
                intf[interface][1] = Line.replace(' description ', '').replace('\n', '')
            elif 'port link-type trunk' in Line and int_flag == 2:
                intf[interface][2] = 'trunk'
            elif re.search(r'^ port trunk allow-pass vlan', Line) and int_flag == 2:
                intf[interface][3] = Line.replace(' port trunk allow-pass vlan ', '').replace('\n', '')
            elif 'port default vlan' in Line and int_flag == 2:
                intf[interface][2] = 'access'
                vid = Line.replace(' port default vlan ', '').replace('\n', '').replace(' ', '')
                intf[interface][3] = int(vid)
                intf[interface][5] = svi[vid]

            elif 'eth-trunk' in Line and int_flag == 2:
                intf[interface][4] = Line.replace(' eth-trunk ', 'eth-trunk').replace('\n', '')
            elif 'shutdown' in Line and int_flag == 2:
                intf[interface][6] = 'ADM'
            elif re.search(r'GigabitEthernet.*/', Line) and int_flag == 3:
                str = Line.replace('GigabitEthernet', 'GE')
                str = str.split()
                int_status[str[0]] = ['', '', '空闲'] #[接口,,占用状态]
                if str[1] == '*down':
                    int_status[str[0]][0] = 'ADM'
                if str[1] != '*down':
                    # print(str)
                    if str[1] == 'up':
                        int_status[str[0]][0] = 'up'
                    else:
                        int_status[str[0]][0] = 'undo shutdown'
                if str[1] == 'up':
                    int_status[str[0]][2] = '占用'
            elif '#' in Line:
                int_flag = 0
            elif 'sysname' in Line:
                HostName = Line.replace('sysname ', '').replace('\n', '')
        return int_status,intf,HostName

FullLogPath = getFullLogPath()
i = 1
wb = Workbook()
ws = wb.active
for x in FullLogPath:
    Fo = open(x, 'r', encoding="UTF-8")
    int_status,intf,HostName = DataProcess(Fo)
    Fo.close()
    i = outPutToexcle(i,int_status,intf,ws,HostName)
wb.save(path_out)
wb.close()
print(Fo)