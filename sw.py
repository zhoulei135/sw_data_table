# -- coding: UTF-8 --
from openpyxl import load_workbook, Workbook
import re

path = 'D:/HuaweiDriver/PyProject/SW/'
path_log = path + 'sw.log'
path_out = path + 'out.xlsx'

fo = open(path_log, 'r', encoding="UTF-8")
#fo = open(path_log, 'r')
wb = Workbook()
ws = wb.active

svi = {}
intf = {}   # [带宽，描述，接口类型，vlan，聚合，ip，shudown]
int_status = {}
int_flag = 0
interface = ''
for line in fo:
    if re.search(r'^interface Vlan-interface', line):   #判断vlan接口
        print(line)
        int_flag = 1
        vlan = line.replace('interface Vlan-interface', '').replace('\n', '')
        svi[vlan] = ''
    elif re.search(r'^interface (Ten-)*GigabitEthernet|^interface HundredGig', line):
        print(line)
        int_flag = 2
        if 'HundredGigE' in line:
            print(line)
            interface = line.replace('interface HundredGigE', 'HGE')
            interface = interface.replace('\n', '')
            intf[interface] = [100, '', '', '', '', '', '']
        if 'Ten-GigabitEthernet' in line:
            print(line)
            interface = line.replace('interface Ten-GigabitEthernet', 'XGE')
            interface = interface.replace('\n', '')
            intf[interface] = [10, '', '', '', '', '', '']
        if ' GigabitEthernet' in line:
            print(line)
            interface = line.replace('interface GigabitEthernet', 'GE')
            interface = interface.replace('\n', '')
            intf[interface] = [1, '', '', '', '', '', '']
    elif 'Type PVID' in line:
        int_flag = 3
    elif 'ip address' in line and int_flag == 1:
        ip = line.split()[2]
        svi[vlan] = ip
    elif ' description' in line and 'level-' not in line and int_flag == 2:
        intf[interface][1] = line.replace(' description ', '').replace('\n', '')
    elif 'link-type trunk' in line and int_flag == 2:
        intf[interface][2] = 'trunk'
    elif re.search(r'^ port trunk permit', line) and int_flag == 2:
        print(line)
        intf[interface][3] = line.replace(' port trunk permit vlan ', '').replace('\n', '')
        print(intf[interface])
    elif 'port access vlan' in line and int_flag == 2:
        intf[interface][2] = 'access'
        vid = line.replace(' port access vlan ', '').replace('\n', '')
        intf[interface][3] = int(vid)
        print(intf[interface][3])
        intf[interface][5] = svi[vid]
    elif 'link-aggregation group' in line and int_flag == 2:
        intf[interface][4] = line.replace(' port link-aggregation group ', 'BAGG').replace('\n', '')
    elif 'shutdown' in line and int_flag == 2:
        intf[interface][6] = 'ADM'
    elif re.search(r'GE\d/', line) and int_flag == 3:
        str = line.split()
        int_status[str[0]] = ['', '', '空闲']
        if str[1] == 'ADM' and (str[4] != 'A' or str[5] != '1' or len(str) == 7):
            print(str)
            int_status[str[0]][0] = 'ADM'
        if str[1] != 'ADM' and str[4] == 'A' and str[5] == '1' and len(str) == 6:
            print(str)
            int_status[str[0]][1] = 'no'
        if  str[1] == 'UP' or str[4] != 'A' or str[5] != '1' or len(str) == 7:
            int_status[str[0]][2] = '占用'
i = 1
ws.cell(row=i, column=4).value = '物理端口'
ws.cell(row=i, column=5).value = '归属逻辑端口'
ws.cell(row=i, column=6).value = '端口类型'
ws.cell(row=i, column=7).value = '端口Vlan'
ws.cell(row=i, column=8).value = '物理带宽(Gbps)'
ws.cell(row=i, column=9).value = '使用状态'
ws.cell(row=i, column=10).value = '本端IP地址'
ws.cell(row=i, column=12).value = '对端设备描述'
i = 2
for k, v in int_status.items():
    ws.cell(row=i, column=4).value = k
    ws.cell(row=i, column=5).value = intf[k][4]
    ws.cell(row=i, column=6).value = intf[k][2]
    ws.cell(row=i, column=7).value = intf[k][3]
    ws.cell(row=i, column=8).value = intf[k][0]
    ws.cell(row=i, column=9).value = v[2]
    ws.cell(row=i, column=10).value = intf[k][5]
    ws.cell(row=i, column=12).value = intf[k][1]
    ws.cell(row=i, column=18).value = v[0]
    ws.cell(row=i, column=19).value = v[1]
    i += 1

#print(svi)
#print(intf)
#print(int_status)

fo.close()
wb.save(path_out)
wb.close()
